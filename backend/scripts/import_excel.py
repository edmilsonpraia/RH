"""
Importa colaboradores reais do ficheiro Excel para a base de dados SQLite.

Uso:
    python backend/scripts/import_excel.py

Lê a folha "CGC  GERAL 2026" (consolidada AHCC + SEDE) e popula
a tabela `employees`. Faz UPSERT por nome (case-insensitive) - re-execucoes
sao idempotentes.

Antes de correr, reinicie o servidor uma vez para garantir que o schema
esta atualizado (ALTER TABLEs em backend/database.js).
"""
import os
import re
import sys
import sqlite3
import unicodedata
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = os.path.join(ROOT, "matriz_competencias_avancada222.xlsx")
DB_PATH = os.path.join(ROOT, "backend", "db", "hr.db")
SHEET_NAME = "CGC  GERAL 2026"


def strip_accents(text: str) -> str:
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())


def make_email(name: str, used: set) -> str:
    """Gera email unico a partir do nome: primeiro.ultimo@copiagroup.com"""
    base = strip_accents(name).lower()
    base = re.sub(r"[^a-z\s]", "", base).strip()
    parts = [p for p in base.split() if p]
    if not parts:
        slug = "colaborador"
    elif len(parts) == 1:
        slug = parts[0]
    else:
        slug = f"{parts[0]}.{parts[-1]}"
    email = f"{slug}@copiagroup.com"
    n = 2
    while email in used:
        email = f"{slug}{n}@copiagroup.com"
        n += 1
    used.add(email)
    return email


def normalize_degree(value):
    if not value:
        return None
    v = strip_accents(str(value)).strip().lower()
    mapping = {
        "tecnico medio": "Técnico Médio",
        "tecnica media": "Técnico Médio",
        "medio": "Técnico Médio",
        "base": "Base",
        "licenciatura": "Licenciatura",
        "licenciado": "Licenciatura",
        "lincenciatura": "Licenciatura",
        "mestre": "Mestrado",
        "pos graduado": "Pós-Graduação",
    }
    return mapping.get(v, str(value).strip())


def normalize_contract(value):
    if not value:
        return None
    v = strip_accents(str(value)).strip().lower()
    if "indeterminado" in v:
        return "Indeterminado"
    if "determinado" in v:
        return "Determinado"
    if "presta" in v:
        return "Prestação de Serviço"
    if "estagi" in v:
        return "Estágio"
    return str(value).strip()


def normalize_gender(value):
    if not value:
        return None
    v = strip_accents(str(value)).strip().lower()
    if v.startswith("masc") or v == "m":
        return "Masculino"
    if v.startswith("fem") or v == "f":
        return "Feminino"
    return str(value).strip()


def normalize_site(value):
    if not value:
        return None
    v = str(value).strip().upper().replace("AIBC_SEDE", "SEDE")
    return v


def normalize_department(value):
    if not value:
        return None
    v = str(value).strip()
    fixes = {
        "Departamento Tecnico": "Departamento Técnico",
        "Pesquisa e Desemvolvimento": "Pesquisa e Desenvolvimento",
        "Projectos Espécias": "Projectos Especiais",
        "Petroléo Gás": "Petróleo e Gás",
        "Seguros": "Seguros",
    }
    return fixes.get(v.strip(), v)


def to_iso_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    try:
        return datetime.fromisoformat(str(val)).strftime("%Y-%m-%d")
    except Exception:
        return None


def to_int(val):
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def main():
    if not os.path.exists(EXCEL_PATH):
        sys.stderr.write(f"Ficheiro Excel nao encontrado: {EXCEL_PATH}\n")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        sys.stderr.write(f"BD nao encontrada: {DB_PATH}\n")
        sys.stderr.write("Inicie o servidor uma vez (`npm start`) antes de importar.\n")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.stderr.write(f"Folha '{SHEET_NAME}' nao encontrada\n")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Carregar emails ja existentes para garantir unicidade
    cur.execute("SELECT email FROM employees")
    used_emails = {r["email"] for r in cur.fetchall() if r["email"]}

    inserted = 0
    updated = 0
    skipped = 0

    for row_idx in range(3, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, 27)]
        (n_seq, meca, nome, doc_tipo, doc_numero, doc_caducidade, doc_estado,
         nacionalidade, idade, nascimento, admissao, contrato_tipo, duracao,
         antiguidade, ult_renov, prox_renov, contrato_estado, departamento,
         funcao, atividade, site, empresa, filiacao, grau, contacto, genero) = cells

        nome = normalize_name(nome)
        if not nome:
            skipped += 1
            continue

        position = to_str(funcao) or "N/D"
        department = normalize_department(departamento) or "N/D"
        hire_date = to_iso_date(admissao) or "1900-01-01"

        # UPSERT: procurar por nome (case-insensitive)
        cur.execute("SELECT id, email FROM employees WHERE LOWER(name) = LOWER(?)", (nome,))
        existing = cur.fetchone()

        data = {
            "meca": to_str(meca),
            "name": nome,
            "phone": to_str(contacto),
            "position": position,
            "department": department,
            "salary": 0,
            "hire_date": hire_date,
            "status": "active",
            "document_type": to_str(doc_tipo),
            "document_number": to_str(doc_numero),
            "document_expiry": to_iso_date(doc_caducidade),
            "document_status": to_str(doc_estado),
            "nationality": to_str(nacionalidade),
            "birth_date": to_iso_date(nascimento),
            "gender": normalize_gender(genero),
            "contract_type": normalize_contract(contrato_tipo),
            "contract_duration_days": to_int(duracao),
            "seniority": to_str(antiguidade),
            "last_renewal_date": to_iso_date(ult_renov),
            "next_renewal_date": to_iso_date(prox_renov),
            "contract_status": to_str(contrato_estado),
            "activity_type": to_str(atividade),
            "site": normalize_site(site),
            "company": to_str(empresa),
            "children": to_int(filiacao) or 0,
            "academic_degree": normalize_degree(grau),
        }

        if existing:
            data["id"] = existing["id"]
            sets = ", ".join([f"{k} = :{k}" for k in data if k != "id"])
            cur.execute(f"UPDATE employees SET {sets}, updated_at = CURRENT_TIMESTAMP WHERE id = :id", data)
            updated += 1
        else:
            data["email"] = make_email(nome, used_emails)
            cols = ", ".join(data.keys())
            placeholders = ", ".join([f":{k}" for k in data.keys()])
            cur.execute(f"INSERT INTO employees ({cols}) VALUES ({placeholders})", data)
            inserted += 1

    conn.commit()

    cur.execute("SELECT COUNT(*) as n FROM employees")
    total = cur.fetchone()["n"]

    cur.execute("SELECT department, COUNT(*) as n FROM employees GROUP BY department ORDER BY n DESC")
    by_dept = cur.fetchall()

    cur.execute("SELECT site, COUNT(*) as n FROM employees GROUP BY site ORDER BY n DESC")
    by_site = cur.fetchall()

    conn.close()

    print(f"Importacao concluida")
    print(f"  Inseridos: {inserted}")
    print(f"  Atualizados: {updated}")
    print(f"  Ignorados: {skipped}")
    print(f"  Total na BD: {total}")
    print()
    print("Por site:")
    for r in by_site:
        print(f"  {r['site'] or '-'}: {r['n']}")
    print()
    print("Por departamento (top 10):")
    for r in by_dept[:10]:
        print(f"  {r['department']}: {r['n']}")


if __name__ == "__main__":
    main()
